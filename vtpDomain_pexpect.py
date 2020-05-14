import datetime
import time
import wexpect
import getpass
import openpyxl
import parse

def createExcelFile():
    # Excel File Creation
    nowDate = 'Report Date: ' + str(datetime.datetime.now().strftime('%Y-%m-%d'))
    wb = openpyxl.load_workbook('VTP_DomainInfo.xlsx')
    ws = wb.active
    ws.title = 'VTP Domain'
    ws['A2'] = nowDate
    ws['A4'] = 'Hostname'
    ws['B4'] = 'IP Address'
    ws['C4'] = 'VTP Domain'
    ws['D4'] = 'VTP Operation Mode'

    wb.save('VTP_DomainInfo.xlsx')
    wb.close()

def saveExcelFile(vtpDomain, vtpOpMode, elem, cellNumber):

    wb = openpyxl.load_workbook('VTP_DomainInfo.xlsx')
    ws = wb.active

    ws['A' + str(cellNumber)] = elem[0]
    ws['B' + str(cellNumber)] = elem[2]
    ws['C' + str(cellNumber)] = vtpDomain['vtpDomain']
    ws['D' + str(cellNumber)] = vtpOpMode['vtpOpMode']

    wb.save('VTP_DomainInfo.xlsx')
    wb.close()

def accessJumpBox(username, password):

    print('\n--- Attempting connection to ' + 'IS6 Server... ')
    ssh_newkey = 'Are you sure you want to continue connecting'
    session = wexpect.spawn('ssh ' + username + '@yourserver ipaddress')

    idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

    if idx == 0:
        session.sendline('yes')
        idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

        if idx == 0:
            session.sendline(password)
    elif idx == 1:
        session.sendline(password)

    idx = session.expect(['$', wexpect.EOF])

    if idx == 0:
        print("--- Successful Login to JumpBox")
        return session
    else:
        print("--- Terminated program")
        exit()

def accessSwitches(session, switch, username, password):

    if 'SSH' in switch[3]:
        ssh_newkey = 'Are you sure you want to continue'
        session.sendline('ssh ' + switch[2])

        print('\n------------------------------------------------------')
        print('--- Attempting connection to: ' + switch[2])
        print('------------------------------------------------------\n')

        idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

        if idx == 0:
            session.sendline('yes')
            time.sleep(.5)
            session.sendline(password)
        elif idx == 1:
            session.sendline(password)
        
    else:
        session.sendline('telnet ' + switch[2])
        
        print('\n------------------------------------------------------')
        print('--- Attempting connection to: ' + switch[2])
        print('------------------------------------------------------\n')

        idx = session.expect(['name', wexpect.EOF])

        if idx == 0:
            session.sendline(username)
            idx = session.expect(['word', wexpect.EOF])
            session.sendline(password)

        else:
            print('Something''s wrong!')
            print('--- Terminated Program!!')
            exit()
    idx = session.expect(['>', '#', wexpect.EOF])
    print('--- Success Login to: ', switch[2])
 
    if idx == 0:
        session.sendline('en')
        idx = session.expect(['word:', wexpect.EOF])
        
    if idx == 0:
        session.sendline(password)
        session.expect(['#', wexpect.EOF])
    
    return session

def getDeviceList():
    deviceList = []

    file = open('0514.txt', 'r')

    for line in file:
        temp = line.split('\t')
        temp[-1] = temp[-1].replace('\n', '')
        deviceList.append(temp)
    file.close()

    return deviceList

def commandExecute(session, os):

    command = ''

    session.sendline('term length 0')
    session.expect(['#', wexpect.EOF])

    if os == 'IOS':
        command += 'sh vtp status'
    elif os == 'NXOS':
        command += 'sh vtp status'
        
    session.sendline(command)
    session.expect(['#', wexpect.EOF])

    return session.before.splitlines()

if __name__ == '__main__':

    cellNumber = 5
    print()
    print('+-------------------------------------------------------------+')
    print('|    Cisco L2 switches VTP Domain Gathernig tool...           |')
    print('|    Version 1.0.0                                            |')
    print('|    Compatible with C35xx, C37xx, C38xx, C65XX               |')
    print('|    Nexus 5K, 7K, 9K                                         |')
    print('|    Scripted by Ethan Park, May. 2020                        |')
    print('+-------------------------------------------------------------+')
    print()
    username = input("Enter your admin ID ==> ")
    password = getpass.getpass("Enter your password ==> ")
    print()

    switchList = getDeviceList()
    createExcelFile()

    for elem in switchList:
        
        session = accessJumpBox(username, password)
        session = accessSwitches(session, elem, username, password)
        data = commandExecute(session, elem[1])
        switch = parse.Parse(data)
        saveExcelFile(switch.getVTPDomain(), switch.getVTPOperationMode(), elem, cellNumber)

        cellNumber += 1
        session.close()